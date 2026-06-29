from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from fpdf import FPDF

from src.account_config import AccountConfig
from src.receipt_metadata_extractor import ReceiptMetadata
from src.receipt_summary import (
    COLUMN_HEADERS_HE,
    COLUMNS,
    _build_row,
    _donor_match_note,
    _find_config_donor_id,
    _parse_amount,
    compute_donor_match,
    compute_severity,
    generate_summary_workbook,
)


def _make_text_pdf(path: Path, text: str) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(text=text)
    pdf.output(str(path))


def _make_empty_pdf(path: Path) -> None:
    pdf = FPDF()
    pdf.add_page()
    pdf.output(str(path))


def _make_corrupt_pdf(path: Path) -> None:
    path.write_bytes(b"not a pdf")


def _setup_receipts(tmp_path: Path, account: str, year: int) -> Path:
    receipts_dir = tmp_path / "receipts"

    may_dir = receipts_dir / account / str(year) / "05_May"
    may_dir.mkdir(parents=True)
    _make_text_pdf(may_dir / "10_05_26__receipt_1001.pdf", "Date: 10/05/2026 Amount $150")
    _make_text_pdf(may_dir / "20_05_26__receipt_1002.pdf", "Date: 20/05/2026 Amount $250")

    jun_dir = receipts_dir / account / str(year) / "06_June"
    jun_dir.mkdir(parents=True)
    _make_text_pdf(jun_dir / "05_06_26__receipt_2001.pdf", "Date: 05/06/2026")
    _make_empty_pdf(jun_dir / "08_06_26__receipt_2002.pdf")
    _make_corrupt_pdf(jun_dir / "10_06_26__receipt_2003.pdf")

    return receipts_dir


def test_workbook_created(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    assert output_path.exists()
    assert output_path.name == "donation_summary_2026.xlsx"


def test_one_sheet_per_populated_month_in_hebrew(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    assert "מאי" in wb.sheetnames
    assert "יוני" in wb.sheetnames
    # English folder names must not appear as sheet names
    assert "05_May" not in wb.sheetnames
    assert "06_June" not in wb.sheetnames


def test_header_row_is_hebrew(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert headers == [COLUMN_HEADERS_HE[col] for col in COLUMNS]


def test_no_file_path_column(tmp_path):
    assert "file_path" not in COLUMNS


def test_file_name_column_present(tmp_path):
    assert "file_name" in COLUMNS
    assert COLUMN_HEADERS_HE["file_name"] == "שם קובץ"


def test_one_row_per_pdf(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    assert wb["מאי"].max_row == 3   # 1 header + 2 PDFs
    assert wb["יוני"].max_row == 4  # 1 header + 3 PDFs


def test_file_name_always_populated(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["יוני"]
    file_name_col = COLUMNS.index("file_name") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=file_name_col).value
        assert cell_value, f"file_name empty on row {row_idx}"


def test_bad_pdfs_become_לבדיקה_and_do_not_abort(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, counts = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    assert output_path.exists()
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["יוני"]
    status_col = COLUMNS.index("extraction_status") + 1
    statuses = [ws.cell(row=r, column=status_col).value for r in range(2, ws.max_row + 1)]
    assert "לבדיקה" in statuses


def test_status_values_are_hebrew(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    status_col = COLUMNS.index("extraction_status") + 1
    all_statuses = set()
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=status_col).value
            if v:
                all_statuses.add(v)
    # All status values in the workbook must be Hebrew, not English
    for s in all_statuses:
        assert s in ("תקין", "חלקי", "לבדיקה"), f"unexpected status: {s}"


def test_counts_use_severity_keys(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    _, counts = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    total = sum(counts.values())
    assert total == 5  # 2 May + 3 June
    # Keys are severity values (used by main.py for printing)
    for key in counts:
        assert key in ("ready", "warning", "critical")


def test_empty_year_creates_סיכום_sheet(tmp_path):
    receipts_dir = tmp_path / "receipts"
    receipts_dir.mkdir(parents=True)
    reports_dir = tmp_path / "reports"

    output_path, counts = generate_summary_workbook(receipts_dir, reports_dir, "emptyaccount", 2026)

    assert output_path.exists()
    assert counts == {}
    wb = openpyxl.load_workbook(str(output_path))
    assert "סיכום" in wb.sheetnames


def test_overwrite_existing_workbook(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)
    output_path2, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    assert output_path2 == output_path
    assert output_path2.exists()


# ---------------------------------------------------------------------------
# account column
# ---------------------------------------------------------------------------


def test_account_column_in_columns():
    assert "account" in COLUMNS


def test_account_column_has_hebrew_header():
    assert COLUMN_HEADERS_HE["account"] == "חשבון"


def test_account_column_value_when_no_config(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    account_col = COLUMNS.index("account") + 1
    assert ws.cell(row=2, column=account_col).value == "testaccount"


def test_account_column_uses_display_name_from_config(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"
    config = AccountConfig(display_name="My Primary Account")

    output_path, _ = generate_summary_workbook(
        receipts_dir, reports_dir, "testaccount", 2026, config=config
    )

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    account_col = COLUMNS.index("account") + 1
    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row=row_idx, column=account_col).value == "My Primary Account"


# ---------------------------------------------------------------------------
# donor columns
# ---------------------------------------------------------------------------


def test_donor_name_column_in_columns():
    assert "donor_name" in COLUMNS


def test_donor_id_column_in_columns():
    assert "donor_id" in COLUMNS


def test_donor_name_column_has_hebrew_header():
    assert COLUMN_HEADERS_HE["donor_name"] == 'שם תורם שזוהה'


def test_donor_id_column_has_hebrew_header():
    assert COLUMN_HEADERS_HE["donor_id"] == 'ת"ז תורם שזוהתה'


def test_donor_columns_appear_in_workbook_header(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert 'שם תורם שזוהה' in headers
    assert 'ת"ז תורם שזוהתה' in headers


def test_account_column_falls_back_to_account_arg_when_config_none(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "myaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(
        receipts_dir, reports_dir, "myaccount", 2026, config=None
    )

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    account_col = COLUMNS.index("account") + 1
    assert ws.cell(row=2, column=account_col).value == "myaccount"


# ---------------------------------------------------------------------------
# compute_donor_match — unit tests (pure function, no PDF needed)
# ---------------------------------------------------------------------------


def _cfg(*, ids=None, names=None) -> AccountConfig:
    return AccountConfig(
        display_name="Test",
        expected_donor_ids=ids or [],
        expected_donor_names=names or [],
    )


def test_donor_match_matched_by_id():
    assert compute_donor_match("123456789", "", _cfg(ids=["123456789"])) == "matched"


def test_donor_match_mismatch_by_id():
    assert compute_donor_match("999999999", "", _cfg(ids=["123456789"])) == "mismatch"


def test_donor_match_id_falls_through_to_name_when_no_id_extracted():
    # expected_donor_ids configured but no donor_id extracted → name check
    assert (
        compute_donor_match(
            "", "ישראל ישראלי", _cfg(ids=["123456789"], names=["ישראל ישראלי"])
        )
        == "matched_by_name"
    )


def test_donor_match_matched_by_name_exact():
    assert compute_donor_match("", "ישראל ישראלי", _cfg(names=["ישראל ישראלי"])) == "matched_by_name"


def test_donor_match_matched_by_name_substring_detected_in_expected():
    # detected name is substring of expected name
    assert compute_donor_match("", "ישראל", _cfg(names=["ישראל ישראלי"])) == "matched_by_name"


def test_donor_match_matched_by_name_expected_in_detected():
    # expected name is substring of detected name
    assert compute_donor_match("", "ישראל ישראלי הגדול", _cfg(names=["ישראל ישראלי"])) == "matched_by_name"


def test_donor_match_matched_by_name_case_insensitive():
    assert compute_donor_match("", "israel israely", _cfg(names=["Israel Israely"])) == "matched_by_name"


def test_donor_match_mismatch_by_name():
    assert compute_donor_match("", "שם אחר", _cfg(names=["ישראל ישראלי"])) == "mismatch"


def test_donor_match_not_detected_no_config():
    assert compute_donor_match("123456789", "Some Name", None) == "not_detected"


def test_donor_match_not_detected_empty_expectation_lists():
    assert compute_donor_match("123456789", "Some Name", _cfg()) == "not_detected"


def test_donor_match_not_detected_no_donor_data_extracted():
    assert compute_donor_match("", "", _cfg(names=["ישראל ישראלי"])) == "not_detected"


# ---------------------------------------------------------------------------
# donor_match column — structure tests
# ---------------------------------------------------------------------------


def test_donor_match_column_in_columns():
    assert "donor_match" in COLUMNS


def test_donor_match_column_has_hebrew_header():
    assert COLUMN_HEADERS_HE["donor_match"] == "התאמת תורם"


def test_donor_match_column_after_donor_id():
    assert COLUMNS.index("donor_match") == COLUMNS.index("donor_id") + 1


# ---------------------------------------------------------------------------
# donor_match column — workbook integration tests
# ---------------------------------------------------------------------------


def test_donor_match_header_appears_in_workbook(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert "התאמת תורם" in headers


def test_donor_match_shows_לא_זוהה_when_no_config(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    col = COLUMNS.index("donor_match") + 1
    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row=row_idx, column=col).value == "לא זוהה"


def test_donor_match_shows_לא_זוהה_when_config_has_expectations_but_no_donor_in_pdf(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"
    config = AccountConfig(display_name="Test", expected_donor_names=["ישראל ישראלי"])

    output_path, _ = generate_summary_workbook(
        receipts_dir, reports_dir, "testaccount", 2026, config=config
    )

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    col = COLUMNS.index("donor_match") + 1
    for row_idx in range(2, ws.max_row + 1):
        assert ws.cell(row=row_idx, column=col).value == "לא זוהה"


def test_donor_match_note_appended_when_expectations_configured_but_no_donor_detected(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"
    config = AccountConfig(display_name="Test", expected_donor_names=["ישראל ישראלי"])

    output_path, _ = generate_summary_workbook(
        receipts_dir, reports_dir, "testaccount", 2026, config=config
    )

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    notes_col = COLUMNS.index("notes") + 1
    notes_values = [ws.cell(row=r, column=notes_col).value or "" for r in range(2, ws.max_row + 1)]
    assert all("לא זוהה תורם" in v for v in notes_values)


def test_donor_match_no_note_when_no_config(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    notes_col = COLUMNS.index("notes") + 1
    notes_values = [ws.cell(row=r, column=notes_col).value or "" for r in range(2, ws.max_row + 1)]
    assert not any("לא זוהה תורם" in v for v in notes_values)


# ---------------------------------------------------------------------------
# compute_severity — unit tests (pure function, no PDF needed)
# ---------------------------------------------------------------------------


def _meta(**kwargs) -> ReceiptMetadata:
    defaults = dict(
        file_name="test.pdf",
        amount="100",
        receipt_date="01/01/2026",
        receipt_number="12345",
        registration_number="123456789",
        tax_report_number="5678",
        organization_name="עמותה",
        donor_name="",
        donor_id="",
        donor_match="not_detected",
    )
    defaults.update(kwargs)
    return ReceiptMetadata(**defaults)


def test_severity_ready_when_all_fields_present_2026():
    assert compute_severity(_meta(), 2026, None) == "ready"


def test_severity_critical_missing_amount():
    assert compute_severity(_meta(amount=""), 2026, None) == "critical"


def test_severity_critical_missing_receipt_date():
    assert compute_severity(_meta(receipt_date=""), 2026, None) == "critical"


def test_severity_critical_missing_receipt_number():
    assert compute_severity(_meta(receipt_number=""), 2026, None) == "critical"


def test_severity_critical_missing_registration_number():
    assert compute_severity(_meta(registration_number=""), 2026, None) == "critical"


def test_severity_critical_missing_tax_report_number_2026():
    assert compute_severity(_meta(tax_report_number=""), 2026, None) == "critical"


def test_severity_ready_missing_tax_report_number_2025():
    assert compute_severity(_meta(tax_report_number=""), 2025, None) == "ready"


def test_severity_critical_donor_mismatch():
    assert compute_severity(_meta(donor_match="mismatch"), 2026, None) == "critical"


def test_severity_warning_donor_not_detected_with_config_expectations():
    config = _cfg(names=["ישראל ישראלי"])
    assert compute_severity(_meta(donor_match="not_detected"), 2026, config) == "warning"


def test_severity_ready_donor_not_detected_no_config():
    assert compute_severity(_meta(donor_match="not_detected"), 2026, None) == "ready"


def test_severity_warning_missing_organization_name():
    assert compute_severity(_meta(organization_name=""), 2026, None) == "warning"


# ---------------------------------------------------------------------------
# severity column — structure tests
# ---------------------------------------------------------------------------


def test_severity_column_in_columns():
    assert "severity" in COLUMNS


def test_severity_column_has_hebrew_header():
    assert COLUMN_HEADERS_HE["severity"] == "חומרה"


def test_severity_column_after_donor_match():
    assert COLUMNS.index("severity") == COLUMNS.index("donor_match") + 1


# ---------------------------------------------------------------------------
# severity column — workbook integration tests
# ---------------------------------------------------------------------------


def test_severity_header_appears_in_workbook(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert "חומרה" in headers


def test_corrupt_pdf_severity_is_critical(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["יוני"]
    severity_col = COLUMNS.index("severity") + 1
    severities = [ws.cell(row=r, column=severity_col).value for r in range(2, ws.max_row + 1)]
    assert "בעיה קריטית" in severities


def test_notes_use_severity_categorised_format(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2026)

    wb = openpyxl.load_workbook(str(output_path))
    notes_col = COLUMNS.index("notes") + 1
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for r in range(2, ws.max_row + 1):
            note = ws.cell(row=r, column=notes_col).value or ""
            for segment in note.split("; "):
                assert not segment.startswith("חסר: "), f"old note format in: {note!r}"


def test_notes_contain_not_required_for_tax_year_when_2025(tmp_path):
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2025)
    reports_dir = tmp_path / "reports"

    output_path, _ = generate_summary_workbook(receipts_dir, reports_dir, "testaccount", 2025)

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb["מאי"]
    notes_col = COLUMNS.index("notes") + 1
    notes = [ws.cell(row=r, column=notes_col).value or "" for r in range(2, ws.max_row + 1)]
    assert all("שדה לא נדרש לשנת מס זו" in v for v in notes)


# --- dual extraction: _best_metadata selection logic ---

from src.hebrew_text_normalizer import normalize_hebrew_text  # noqa: E402
from src.receipt_metadata_extractor import extract_metadata  # noqa: E402
from src.receipt_summary import _best_metadata  # noqa: E402

_TRANZILA_PATH = Path("receipts/primary/2026/06_June/24_06_26__tranzila_test.pdf")
_TRANZILA_RAW = (
    "לכבוד\n"
    "אוריאל אוחיון\n"
    'ת"ז 111111118\n'
    'מלכ"ר : 580537942\n'
    "קבלה תרומה 302678\n"
    "אישור דיווח: 88287\n"
    "24/06/2026\n"
    "₪100"
)
# Normalizing logical-order Hebrew reverses anchors → breaks receipt/tax/donor_name patterns
_TRANZILA_NORM = normalize_hebrew_text(_TRANZILA_RAW)


def test_best_metadata_selects_raw_when_raw_wins():
    meta = _best_metadata(_TRANZILA_PATH, _TRANZILA_RAW, _TRANZILA_NORM)
    # These three fields are only recoverable from raw (anchors broken in normalized)
    assert meta.receipt_number == "302678"
    assert meta.tax_report_number == "88287"
    assert meta.donor_name == "אוריאל אוחיון"


def test_best_metadata_selects_normalized_when_normalized_wins():
    # Sparse raw (2 fields); richer synthetic normalized (4 fields) → normalized chosen
    sparse_raw = "24/06/2026\n₪100"
    richer_norm = (
        "עמותה\n580537942\n"
        "קבלה תרומה 302678\n"
        "24/06/2026\n₪100"
    )
    meta = _best_metadata(_TRANZILA_PATH, sparse_raw, richer_norm)
    assert meta.registration_number == "580537942"
    assert meta.receipt_number == "302678"


def test_best_metadata_tie_prefers_normalized():
    # Same fields (date + amount) but different amounts → normalized wins for amount field
    raw_text = "24/06/2026\n₪100"
    norm_text = "24/06/2026\n₪200"
    path = Path("receipts/primary/2026/06_June/24_06_26__tie_test.pdf")
    meta = _best_metadata(path, raw_text, norm_text)
    assert meta.amount == "200"


# --- field-level merge ---


def test_best_metadata_merge_raw_registration_with_norm_structural_fields():
    # raw: only registration anchor; norm: only receipt/date/amount.
    # Old whole-object selection would lose registration_number when norm has more fields.
    # New merge must include all fields from both sources.
    raw = "עמותה רשומה: 580712348"
    norm = "קבלה תרומה 80806\n05/06/2026\n₪200.00"
    path = Path("receipts/primary/2026/06_June/receipt.pdf")
    meta = _best_metadata(path, raw, norm)
    assert meta.registration_number == "580712348"
    assert meta.receipt_number == "80806"
    assert meta.receipt_date == "05/06/2026"
    assert meta.amount == "200.00"


def test_best_metadata_merge_raw_donor_fields_preserved():
    # Normalized text reverses Hebrew anchors → breaks donor_name/donor_id extraction.
    # Field-level merge must recover them from raw.
    raw = 'שם התורם: ישראל ישראלי\nת"ז: 123456789\n05/06/2026\n₪100'
    norm = normalize_hebrew_text(raw)
    path = Path("receipts/primary/2026/06_June/05_06_26__receipt_1.pdf")
    meta = _best_metadata(path, raw, norm)
    assert meta.donor_name == "ישראל ישראלי"
    assert meta.donor_id == "123456789"


def test_best_metadata_notes_no_duplicates():
    # When both sources produce the same note (e.g. filename-date note), it must
    # appear only once in the merged metadata.
    text = "עמותה\n580395051\n₪200"  # no in-text date → both sources add filename-date note
    path = Path("receipts/primary/2026/06_June/05_06_26__receipt.pdf")
    meta = _best_metadata(path, text, text)
    assert "תאריך מתוך שם הקובץ" in meta.notes
    assert meta.notes.count("תאריך מתוך שם הקובץ") == 1


# ---------------------------------------------------------------------------
# _parse_amount — unit tests
# ---------------------------------------------------------------------------


def test_parse_amount_bare_integer():
    assert _parse_amount("200") == 200.0


def test_parse_amount_decimal():
    assert _parse_amount("200.00") == 200.0


def test_parse_amount_shekel_prefix():
    assert _parse_amount("₪200") == 200.0


def test_parse_amount_shekel_with_space():
    assert _parse_amount("₪ 200.00") == 200.0


def test_parse_amount_thousands_separator():
    assert _parse_amount("1,200") == 1200.0


def test_parse_amount_thousands_with_decimal():
    assert _parse_amount("1,200.50") == 1200.5


def test_parse_amount_empty_returns_none():
    assert _parse_amount("") is None


def test_parse_amount_invalid_returns_none():
    assert _parse_amount("N/A") is None


# ---------------------------------------------------------------------------
# donor match — matched_by_name
# ---------------------------------------------------------------------------


def test_compute_donor_match_id_matched_is_still_matched():
    assert compute_donor_match("123456789", "", _cfg(ids=["123456789"])) == "matched"


def test_donor_match_note_matched_by_name_returns_note():
    note = _donor_match_note("", "ישראל ישראלי", None, "matched_by_name")
    assert note == 'תורם תואם לפי שם; ת"ז לא זוהתה בקבלה'


def test_donor_match_note_matched_by_id_returns_empty():
    assert _donor_match_note("123456789", "", None, "matched") == ""


def test_compute_severity_matched_by_name_is_ready():
    config = _cfg(names=["ישראל ישראלי"])
    m = _meta(donor_match="matched_by_name")
    assert compute_severity(m, 2026, config) == "ready"


# ---------------------------------------------------------------------------
# _find_config_donor_id — unit tests
# ---------------------------------------------------------------------------


def test_find_config_donor_id_matches_without_anchor():
    text = "לכבוד\nישראל ישראלי\n111111118\nקבלה 100001\n₪200"
    assert _find_config_donor_id(text, ["111111118"]) == "111111118"


def test_find_config_donor_id_id_not_in_config_returns_none():
    # ID present in text but not listed in config — must not match
    text = "111111118"
    assert _find_config_donor_id(text, ["999999999"]) is None


def test_find_config_donor_id_empty_config_returns_none():
    assert _find_config_donor_id("111111118", []) is None


def test_find_config_donor_id_no_partial_match():
    # "11111111" (8 digits) is a prefix of "111111118" — word boundary must prevent match
    assert _find_config_donor_id("111111118", ["11111111"]) is None


def test_find_config_donor_id_returns_first_matching():
    text = "111111118 and 222222226"
    assert _find_config_donor_id(text, ["222222226", "111111118"]) == "222222226"


def test_donor_match_matched_by_name_displays_as_התאמה_in_workbook(tmp_path):
    # Verify matched_by_name is mapped to the same Hebrew display value as matched
    receipts_dir = _setup_receipts(tmp_path, "testaccount", 2026)
    reports_dir = tmp_path / "reports"
    # The fixture PDFs have no donor data, so donor_match will be not_detected.
    # We just verify the mapping dict value directly via the already-imported constant.
    from src.receipt_summary import _DONOR_MATCH_HE  # noqa: PLC0415
    assert _DONOR_MATCH_HE["matched_by_name"] == "התאמה"
    assert _DONOR_MATCH_HE["matched"] == "התאמה"


# ---------------------------------------------------------------------------
# donor_id completion from config — name-based fallback
# ---------------------------------------------------------------------------


_DONOR_ID_COMPLETION_TEXT = (
    "לכבוד\n"
    "Israel Israely\n"
    "580537942\n"
    "05/06/2026\n"
    "₪100"
)


def test_donor_id_completed_from_config_when_single_id_and_name_matches(tmp_path):
    """When exactly one expected_donor_id is configured and the donor name from the
    receipt matches expected_donor_names, _build_row must populate donor_id from config
    and add the completion note."""
    pdf_path = tmp_path / "05_06_26__receipt_9001.pdf"
    pdf_path.write_bytes(b"fake")

    config = AccountConfig(
        display_name="Test",
        expected_donor_ids=["111111118"],
        expected_donor_names=["Israel Israely"],
    )

    with (
        patch("src.receipt_summary.extract_text_from_pdf", return_value=_DONOR_ID_COMPLETION_TEXT),
        patch("src.receipt_summary.normalize_hebrew_text", return_value=_DONOR_ID_COMPLETION_TEXT),
    ):
        meta = _build_row(pdf_path, "testaccount", config, 2026)

    assert meta.donor_id == "111111118"
    assert 'תעודת זהות הושלמה לפי שם תורם תואם' in meta.notes


def test_donor_id_not_guessed_when_multiple_expected_ids_and_name_only_matches(tmp_path):
    """When config has multiple expected_donor_ids and the donor ID is not found in text,
    _build_row must NOT guess — donor_id must remain empty even if the name matches."""
    pdf_path = tmp_path / "05_06_26__receipt_9002.pdf"
    pdf_path.write_bytes(b"fake")

    config = AccountConfig(
        display_name="Test",
        expected_donor_ids=["111111118", "222222226"],
        expected_donor_names=["Israel Israely"],
    )

    with (
        patch("src.receipt_summary.extract_text_from_pdf", return_value=_DONOR_ID_COMPLETION_TEXT),
        patch("src.receipt_summary.normalize_hebrew_text", return_value=_DONOR_ID_COMPLETION_TEXT),
    ):
        meta = _build_row(pdf_path, "testaccount", config, 2026)

    assert meta.donor_id == ""
    assert meta.donor_match == "matched_by_name"
