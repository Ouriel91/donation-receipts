from __future__ import annotations

from pathlib import Path
import re

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.account_config import AccountConfig
from src.attachment_saver import MONTH_NAMES
from src.hebrew_text_normalizer import normalize_hebrew_text
from src.pdf_text_extractor import (
    PdfExtractionError,
    extract_text_from_pdf,
    extract_text_from_pdf_pymupdf,
)
from src.receipt_metadata_extractor import (
    STATUS_NEEDS_REVIEW,
    STATUS_OK,
    STATUS_PARTIAL,
    ReceiptMetadata,
    extract_metadata,
)

_AMOUNT_STRIP_RE = re.compile(r"[₪\s,]")


def _parse_amount(value: str) -> float | None:
    if not value:
        return None
    cleaned = _AMOUNT_STRIP_RE.sub("", value)
    try:
        return float(cleaned)
    except ValueError:
        return None


def _find_config_donor_id(text: str, expected_ids: list[str]) -> str | None:
    for id_ in expected_ids:
        if re.search(rf"\b{re.escape(id_)}\b", text):
            return id_
    return None


def _donor_name_matches(donor_name: str, expected_names: list[str]) -> bool:
    dn = donor_name.lower()
    return any(dn in e.lower() or e.lower() in dn for e in expected_names)



def _score_metadata(meta: ReceiptMetadata) -> int:
    """Count non-empty critical fields to compare extraction quality across engines."""
    fields = ("amount", "receipt_date", "receipt_number", "registration_number", "tax_report_number")
    return sum(1 for f in fields if getattr(meta, f))


def _best_metadata(pdf_path: Path, raw_text: str, normalized_text: str) -> ReceiptMetadata:
    meta_raw = extract_metadata(raw_text, pdf_path)
    meta_norm = extract_metadata(normalized_text, pdf_path)

    def _pick(*vals: str) -> str:
        return next((v for v in vals if v), "")

    # Field-level merge: raw preserves correct Hebrew anchors (registration, donor);
    # normalized handles fragmented RTL layouts better (receipt/tax/date/amount).
    merged = ReceiptMetadata(
        file_name=meta_raw.file_name,
        registration_number=_pick(meta_raw.registration_number, meta_norm.registration_number),
        donor_name=_pick(meta_raw.donor_name, meta_norm.donor_name),
        donor_id=_pick(meta_raw.donor_id, meta_norm.donor_id),
        receipt_number=_pick(meta_norm.receipt_number, meta_raw.receipt_number),
        tax_report_number=_pick(meta_norm.tax_report_number, meta_raw.tax_report_number),
        receipt_date=_pick(meta_norm.receipt_date, meta_raw.receipt_date),
        amount=_pick(meta_norm.amount, meta_raw.amount),
        organization_name=_pick(meta_norm.organization_name, meta_raw.organization_name),
    )

    # Recompute extraction_status against merged fields.
    missing_critical: list[str] = []
    if not merged.amount:
        missing_critical.append("סכום")
    if not merged.receipt_date:
        missing_critical.append("תאריך")
    if not merged.organization_name and not merged.registration_number:
        missing_critical.append("שם עמותה")

    missing_noncritical: list[str] = []
    if not merged.organization_name and merged.registration_number:
        missing_noncritical.append("שם עמותה")
    if not merged.registration_number:
        missing_noncritical.append("מספר עמותה")
    if not merged.receipt_number:
        missing_noncritical.append("מספר קבלה")
    if not merged.tax_report_number:
        missing_noncritical.append("מספר אישור דיווח")

    if missing_critical:
        merged.extraction_status = STATUS_NEEDS_REVIEW
        merged.notes = f"חסר: {', '.join(missing_critical)}"
    elif missing_noncritical:
        merged.extraction_status = STATUS_PARTIAL
        merged.notes = ""
    else:
        merged.extraction_status = STATUS_OK
        merged.notes = ""

    # Combine non-"חסר:" notes from both sources (e.g. "תאריך מתוך שם הקובץ"), deduplicated.
    surviving: list[str] = []
    seen: set[str] = set()
    for note in (meta_raw.notes + "; " + meta_norm.notes).split("; "):
        n = note.strip()
        if n and not n.startswith("חסר:") and n not in seen:
            seen.add(n)
            surviving.append(n)
    if surviving:
        merged.notes = "; ".join(filter(None, [merged.notes] + surviving))

    return merged

# Internal field names — used for getattr(meta, col) and as dict keys.
COLUMNS = [
    "account",
    "file_name",
    "organization_name",
    "registration_number",
    "receipt_number",
    "tax_report_number",
    "receipt_date",
    "amount",
    "donor_name",
    "donor_id",
    "donor_match",
    "severity",
    "extraction_status",
    "notes",
]

COLUMN_HEADERS_HE: dict[str, str] = {
    "account": "חשבון",
    "file_name": "שם קובץ",
    "organization_name": "שם עמותה",
    "registration_number": "מספר עמותה",
    "receipt_number": "מספר קבלה",
    "tax_report_number": "מספר אישור דיווח",
    "receipt_date": "תאריך",
    "amount": "סכום",
    "donor_name": 'שם תורם שזוהה',
    "donor_id": 'ת"ז תורם שזוהתה',
    "donor_match": "התאמת תורם",
    "severity": "חומרה",
    "extraction_status": "סטטוס",
    "notes": "הערות",
}

_STATUS_HE: dict[str, str] = {
    "ok": "תקין",
    "partial": "חלקי",
    "needs_review": "לבדיקה",
}

_DONOR_MATCH_HE: dict[str, str] = {
    "matched": "התאמה",
    "matched_by_name": "התאמה",
    "not_detected": "לא זוהה",
    "mismatch": "אי התאמה",
}

_SEVERITY_HE: dict[str, str] = {
    "ready": "מוכן",
    "warning": "דורש תשומת לב",
    "critical": "בעיה קריטית",
}

_CRITICAL_FIELDS_HE: dict[str, str] = {
    "amount": "סכום",
    "receipt_date": "תאריך",
    "receipt_number": "מספר קבלה",
    "registration_number": "מספר עמותה",
    "tax_report_number": "מספר אישור דיווח",
}

_MONTH_SHEET_NAMES: dict[str, str] = {
    "01_January": "ינואר",
    "02_February": "פברואר",
    "03_March": "מרץ",
    "04_April": "אפריל",
    "05_May": "מאי",
    "06_June": "יוני",
    "07_July": "יולי",
    "08_August": "אוגוסט",
    "09_September": "ספטמבר",
    "10_October": "אוקטובר",
    "11_November": "נובמבר",
    "12_December": "דצמבר",
}


def iter_month_pdf_files(receipts_dir: Path, account: str, year: int):
    """Yield (month_folder_name, sorted pdf paths) for each month that has PDFs."""
    year_dir = receipts_dir / account / str(year)
    for month_num in sorted(MONTH_NAMES):
        month_folder = year_dir / MONTH_NAMES[month_num]
        if not month_folder.is_dir():
            continue
        pdfs = sorted(month_folder.glob("*.pdf"))
        if pdfs:
            yield MONTH_NAMES[month_num], pdfs


def _build_row(
    pdf_path: Path,
    account: str,
    config: AccountConfig | None,
    year: int,
) -> ReceiptMetadata:
    try:
        raw_text = extract_text_from_pdf(pdf_path)
        normalized = normalize_hebrew_text(raw_text)
        meta = _best_metadata(pdf_path, raw_text, normalized)

        raw_mupdf = extract_text_from_pdf_pymupdf(pdf_path)
        if raw_mupdf:
            norm_mupdf = normalize_hebrew_text(raw_mupdf)
            meta_mupdf = _best_metadata(pdf_path, raw_mupdf, norm_mupdf)
            if _score_metadata(meta_mupdf) > _score_metadata(meta):
                meta, raw_text, normalized = meta_mupdf, raw_mupdf, norm_mupdf

        if not meta.donor_id and config and config.expected_donor_ids:
            found = _find_config_donor_id(raw_text, config.expected_donor_ids)
            if not found:
                found = _find_config_donor_id(normalized, config.expected_donor_ids)
            if found:
                meta.donor_id = found
                if meta.donor_id == meta.registration_number:
                    meta.registration_number = ""
            elif (
                len(config.expected_donor_ids) == 1
                and config.expected_donor_names
                and meta.donor_name
                and _donor_name_matches(meta.donor_name, config.expected_donor_names)
            ):
                meta.donor_id = config.expected_donor_ids[0]
                if meta.donor_id == meta.registration_number:
                    meta.registration_number = ""
                meta.notes = "; ".join(
                    filter(None, [meta.notes, 'תעודת זהות הושלמה לפי שם תורם תואם'])
                )
    except Exception:
        meta = ReceiptMetadata(
            file_name=pdf_path.name,
            extraction_status=STATUS_NEEDS_REVIEW,
            notes="לא ניתן לקרוא את הקובץ",
        )
    meta.account = config.display_name if config else account
    meta.donor_match = compute_donor_match(meta.donor_id, meta.donor_name, config)
    meta.severity = compute_severity(meta, year, config)

    # Replace generic "חסר: ..." extraction notes with severity-categorised equivalents.
    surviving_notes = [n for n in meta.notes.split("; ") if n and not n.startswith("חסר: ")]
    field_notes = _severity_field_notes(meta, year)
    donor_note = _donor_match_note(meta.donor_id, meta.donor_name, config, meta.donor_match)
    meta.notes = "; ".join(filter(None, field_notes + surviving_notes + [donor_note]))

    return meta


def compute_donor_match(
    donor_id: str,
    donor_name: str,
    config: AccountConfig | None,
) -> str:
    if config and config.expected_donor_ids:
        if donor_id and donor_id in config.expected_donor_ids:
            return "matched"
        if donor_id:
            return "mismatch"
        # donor_id not extracted — fall through to name check
    if config and config.expected_donor_names:
        if donor_name:
            dn = donor_name.lower()
            for expected in config.expected_donor_names:
                if dn in expected.lower() or expected.lower() in dn:
                    return "matched_by_name"
            return "mismatch"
    return "not_detected"


def _donor_match_note(
    donor_id: str,
    donor_name: str,  # noqa: ARG001
    config: AccountConfig | None,
    match: str,
) -> str:
    if match == "matched":
        return ""
    if match == "matched_by_name":
        return 'תורם תואם לפי שם; ת"ז לא זוהתה בקבלה'
    if match == "mismatch":
        if config and config.expected_donor_ids and donor_id:
            return "תעודת זהות תורם לא תואמת"
        return "שם תורם לא תואם"
    # not_detected
    if config and (config.expected_donor_ids or config.expected_donor_names):
        return "לא זוהה תורם"
    return ""


def compute_severity(
    meta: ReceiptMetadata,
    year: int,
    config: AccountConfig | None,
) -> str:
    if not meta.amount or not meta.receipt_date or not meta.receipt_number or not meta.registration_number:
        return "critical"
    if year >= 2026 and not meta.tax_report_number:
        return "critical"
    if meta.donor_match == "mismatch":
        return "critical"
    if (
        meta.donor_match == "not_detected"
        and config
        and (config.expected_donor_ids or config.expected_donor_names)
    ):
        return "warning"
    if not meta.organization_name:
        return "warning"
    return "ready"


def _severity_field_notes(meta: ReceiptMetadata, year: int) -> list[str]:
    notes: list[str] = []
    for field in ("amount", "receipt_date", "receipt_number", "registration_number"):
        if not getattr(meta, field):
            notes.append(f"שדה חובה חסר: {_CRITICAL_FIELDS_HE[field]}")
    if not meta.tax_report_number:
        if year >= 2026:
            notes.append(f"שדה חובה חסר: {_CRITICAL_FIELDS_HE['tax_report_number']}")
        else:
            notes.append("שדה לא נדרש לשנת מס זו")
    if not meta.organization_name and meta.registration_number:
        notes.append("שדה אופציונלי חסר: שם עמותה")
    return notes


def _cell_value(col: str, meta: ReceiptMetadata) -> object:
    value = getattr(meta, col)
    if col == "extraction_status":
        return _STATUS_HE.get(value, value)
    if col == "donor_match":
        return _DONOR_MATCH_HE.get(value, value)
    if col == "severity":
        return _SEVERITY_HE.get(value, value)
    if col == "amount":
        numeric = _parse_amount(value)
        return numeric if numeric is not None else value
    return value


def _write_month_sheet(
    wb: openpyxl.Workbook,
    month_folder: str,
    pdfs: list[Path],
    account: str,
    config: AccountConfig | None,
    year: int,
) -> dict[str, int]:
    sheet_title = _MONTH_SHEET_NAMES.get(month_folder, month_folder)
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_HEADERS_HE[col])
        cell.font = header_font

    counts: dict[str, int] = {}
    for row_idx, pdf_path in enumerate(pdfs, start=2):
        meta = _build_row(pdf_path, account, config, year)
        counts[meta.severity] = counts.get(meta.severity, 0) + 1
        for col_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(col, meta))
            if col == "amount":
                cell.number_format = "#,##0.##"

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

    return counts


def _write_empty_summary_sheet(wb: openpyxl.Workbook, account: str, year: int) -> None:
    ws = wb.create_sheet(title="סיכום")
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value=f"סיכום תרומות — {account} / {year}")
    ws.cell(row=2, column=1, value="לא נמצאו קבלות לשנה זו.")


def generate_summary_workbook(
    receipts_dir: Path,
    reports_dir: Path,
    account: str,
    year: int,
    config: AccountConfig | None = None,
) -> tuple[Path, dict[str, int]]:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_counts: dict[str, int] = {}
    month_entries = list(iter_month_pdf_files(receipts_dir, account, year))

    if not month_entries:
        _write_empty_summary_sheet(wb, account, year)
    else:
        for month_folder, pdfs in month_entries:
            counts = _write_month_sheet(wb, month_folder, pdfs, account, config, year)
            for status, n in counts.items():
                total_counts[status] = total_counts.get(status, 0) + n

    output_path = reports_dir / account / str(year) / f"donation_summary_{year}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path, total_counts
